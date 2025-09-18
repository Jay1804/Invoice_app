import os
import re
import pandas as pd
import streamlit as st
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
from num2words import num2words

st.title("Invoice Generator App")

# File uploads
consultant_file = st.file_uploader("Upload Consultant Excel File", type=["xlsx"])
invoice_template = st.file_uploader("Upload Invoice Template", type=["xlsx"])
output_folder = st.text_input("Output Folder Path", r"C:\Users\jay.chaudhary\Desktop\Invoice\Invoice hub")

# Create output folder if not exists
if output_folder:
    os.makedirs(output_folder, exist_ok=True)

# Helper function to safely write to merged cells
def safe_write(ws, cell, value):
    target = ws[cell]
    if target.__class__.__name__ == "MergedCell":
        for merged_range in ws.merged_cells.ranges:
            if cell in merged_range:
                top_left = ws.cell(
                    row=merged_range.min_row,
                    column=merged_range.min_col
                ).coordinate
                ws[top_left] = value
                return
    else:
        ws[cell] = value

# Generate invoices when button is clicked
if st.button("Generate Invoices"):
    if not consultant_file or not invoice_template or not output_folder:
        st.error("Please upload both files and specify an output folder.")
    else:
        # Load consultant Excel
        consultants_df = pd.read_excel(consultant_file)
        st.write(f"Columns detected: {consultants_df.columns.tolist()}")

        generated_files = []

        for idx, row in consultants_df.iterrows():
            # Load template workbook
            wb = load_workbook(invoice_template)
            ws = wb.active

            # Invoice date formatting
            invoice_date = row.get("InvoiceDate", "")
            if pd.notnull(invoice_date):
                try:
                    invoice_date = pd.to_datetime(invoice_date).strftime("%d %b-%y")
                except Exception:
                    invoice_date = str(invoice_date)

            # Invoice number
            raw_invoice_no = str(row.get("InvoiceNo", "")).strip()
            final_invoice_no = f"Invoice No : {raw_invoice_no}"

            # Mandatory fields
            safe_write(ws, "A1", f"Name : {row.get('EMPLOYEE NAME', '')}")
            safe_write(ws, "A2", f"Address : {row.get('Address', '')}")
            safe_write(ws, "A4", f"PAN : {row.get('PAN', '')}")
            safe_write(ws, "A9", final_invoice_no)
            safe_write(ws, "A10", f"Invoice Date : {invoice_date}")
            safe_write(ws, "A12", f"State : {row.get('State', '')}")

            # Specific cells
            safe_write(ws, "H12", row.get("Code", ""))

            # IN HAND amount
            in_hand = row.get("IN HAND", "")
            if pd.notnull(in_hand) and str(in_hand).strip() != "":
                try:
                    in_hand = float(in_hand)
                except Exception:
                    in_hand = None

            safe_write(ws, "N21", in_hand)
            if in_hand is not None:
                ws["N21"].number_format = '#,##0.00'

                rupees = int(in_hand)
                paise = int(round((in_hand - rupees) * 100))

                if paise > 0:
                    in_hand_words = f"{num2words(rupees, lang='en').title()} Rupees And {num2words(paise, lang='en').title()} Paise Only"
                else:
                    in_hand_words = f"{num2words(rupees, lang='en').title()} Rupees Only"

                safe_write(ws, "A48", in_hand_words)
                ws["A48"].font = Font(bold=True, size=20)
                ws["A48"].alignment = Alignment(horizontal='center', vertical='center')

            # Bank details
            safe_write(ws, "B53", row.get("Bankname", ""))
            safe_write(ws, "B54", row.get("Name", ""))
            safe_write(ws, "B55", row.get("AccountNo", ""))
            safe_write(ws, "B56", row.get("IFSC", ""))
            safe_write(ws, "J55", row.get("EMPLOYEE NAME", ""))

            # Clean filename
            emp_name_clean = re.sub(r'[<>:"/\\|?*]', '-', str(row.get("EMPLOYEE NAME", "Unknown")))
            invoice_no_clean = re.sub(r'[<>:"/\\|?*]', '-', str(row.get("InvoiceNo", "")))

            # Save invoice
            output_file = os.path.join(output_folder, f"Invoice_{emp_name_clean}_{invoice_no_clean}.xlsx")
            wb.save(output_file)
            generated_files.append(output_file)

        st.success(f"🎉 {len(generated_files)} invoices generated successfully!")
        st.write("Generated files:")
        for f in generated_files:
            st.write(f)
